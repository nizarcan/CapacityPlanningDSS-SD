import pandas as pd
# import numpy as np
from openpyxl import *
from app_backend.finder import np as np
import app_backend.finder as finder


def level_lookup(df, level_col, lookup_col):
    dummies = pd.get_dummies(df[level_col])

    idx = dummies.index.to_series()
    last_index = dummies.apply(lambda col: idx.where(col != 0, np.nan).fillna(method = "ffill"))
    last_index[0] = 1

    idx = last_index.lookup(last_index.index, df[level_col] - 1)
    return pd.DataFrame({lookup_col: df.reindex(idx)[lookup_col].values}, index = df.index)


def level_lookup_bak(df_name, level_col, lookup_col, level_diff):
    # Creating a dummy matrix with the levels of the bill of materials.
    dummies = pd.get_dummies(df_name[level_col])

    idx = dummies.index.to_series()
    last_index = dummies.apply(lambda col: idx.where(col != 0, np.nan).fillna(method = "ffill"))
    last_index[0] = np.nan

    idx = last_index.lookup(last_index.index, df_name[level_col] - level_diff)
    return pd.DataFrame({lookup_col: df_name.reindex(idx)[lookup_col].values for i in range(len(lookup_col))},
                        index = df_name.index)


def create_legend_table(df):
    df = df.iloc[finder.input_indices(df)][["product_no", "part_no"]]
    df.index = list(range(1, len(df) + 1))
    return df


def create_input_table(df):
    df = df.iloc[finder.input_indices(df)][["product_no", "amount"]]
    df["product_no"] = finder.product_numerator(df)
    df.index = list(range(1, len(df) + 1))
    df.columns = ["product", "amount"]
    return df


def create_duplication_table(df):
    # This function gets the input table as a base dataframe to work on and make calculations with.
    df = create_input_table(df)

    # The following three lines creates the products' index in the process input list, i.e. from the input table
    s = df["product"].ne(df["product"].shift(fill_value = df.iloc[0]["product"]))
    product_idx = pd.Series([1] + list(np.where(s)[0] + 1))
    product_idx.index += 1

    # Following line calculates the entity amounts to be duplicated in the simulation software
    dup_count = product_idx.shift(-1, fill_value = len(df) + 1) - product_idx

    # The next two lines concatanates, basically zipps the created product index and the duplication amounts and
    # converts them to a pandas dataframe with the product # with them.
    duplication_table = pd.concat(
        [pd.Series(list(range(1, len(product_idx) + 1)), index = list(range(1, len(product_idx) + 1))), product_idx,
         dup_count], axis = 1)
    duplication_table.columns = ["product", "start", "number to duplicate"]
    return duplication_table


def find_product_codes(df):
    return df["product_no"].ne(df["product_no"].shift(1, fill_value = df.iloc[0]["product_no"])).cumsum() + 1


def find_process_input_indices(df):
    s = df["level"].ge(df["level"].shift(-1, fill_value = df.iloc[-1]["level"]))
    return np.where(s)


def get_excel_sheet_names(directory):
    wb = load_workbook(directory, read_only = True)
    return wb.sheetnames


def read_excel_sheet(directory, name_of_sheet="no_name"):
    if name_of_sheet == "no_name":
        wb = load_workbook(directory, read_only = True)
        sheet = wb[wb.sheetnames[0]]
        df = pd.DataFrame(sheet.values)
        df.columns = df.iloc[0, :]
        df = df.drop(0)
        df.reset_index(drop = "index", inplace = True)
        return df
    wb = load_workbook(directory, read_only = True)
    sheet = wb[name_of_sheet]
    df = pd.DataFrame(sheet.values)
    df.columns = df.iloc[0, :]
    df = df.drop(0)
    df.reset_index(drop = "index", inplace = True)
    return df


def determine_amounts(df):
    df["new_amount"] = 1
    bom_row = 0
    while bom_row < len(df):
        if (df.loc[bom_row, "amount"]%1 == 0) and (df.loc[bom_row, "amount"] != 1):
            starting_amount = df.loc[bom_row, "amount"]
            df.loc[bom_row, "new_amount"] = starting_amount
            starting_level = df.loc[bom_row, "level"]
            while bom_row < len(df):
                bom_row += 1
                if df.at[bom_row, "level"] > starting_level:
                    df.at[bom_row, "new_amount"] = starting_amount
                else:
                    break
        bom_row += 1
    df['amount'] = df['new_amount']
    df.drop('new_amount', axis = 1, inplace = True)
    df.reset_index(drop = "index", inplace = True)
    return df


def reformat_columns(df, relevant_col_idx, df_type):
    if df_type == "bom":
        # Rearranging the level amount
        df["Seviye"] = [int(x[-1]) for x in df[df.columns[relevant_col_idx][2]]]
        relevant_col_idx[2] = len(df.columns) - 1

        # Determining the columns names to use for reindex later
        relevant_col_names = df.columns[relevant_col_idx]

        # Columns to be dropped from the dataframe
        cols_to_drop = list(set(df.columns) - set(df.columns[relevant_col_idx]))

        # Dropping, sorting and indexing the corresponding columns
        df.drop(cols_to_drop, axis = 1, inplace = True)
        df = df.reindex(columns = relevant_col_names)
        df.columns = ["product_no", "part_no", "level", "amount", "explanation"]

    if df_type == "times":
        # Determining the columns names to use for reindex later
        relevant_col_names = df.columns[relevant_col_idx]

        # Columns to be dropped from the dataframe
        cols_to_drop = list(set(df.columns) - set(df.columns[relevant_col_idx]))

        # Dropping, sorting and indexing the corresponding columns
        df.drop(cols_to_drop, axis = 1, inplace = True)
        df = df.reindex(columns = relevant_col_names)
        df.columns = ["part_no", "station", "cycle_times", "prep_times"]

    return df


def arrange_df(df, relevant_col_idx, df_type, items_to_delete_dir="no_dir", assembly_df=None, cmy_df=None):
    if df_type.lower() == "bom":
        # Reformatting the columns
        df = reformat_columns(df, relevant_col_idx, "bom")

        # This to be deleted parts can be redundant, so it will be decided that if these codes are going to stay or not
        items_to_delete = read_excel_sheet(items_to_delete_dir)
        s = [str(x).split(".")[0] not in list(items_to_delete["Kalacaklar"]) for x in df["part_no"]]
        df = df.drop(np.where(np.asarray(s))[0])

        # Deleting the entries where two successive entries are level 1s.
        s = (df[df.columns[2]].ne(df[df.columns[2]].shift(1, fill_value = 1)) | df[df.columns[2]].ne(1)).cumsum()
        df = df.groupby(s).tail(1)

        # This is the part where the deletion occurs if the last column's level is 1.
        if df.at[df.index[-1], df.columns[2]] == 1:
            df.drop(df.index[-1], inplace = True)

        # Making sure that the dataframe returns in order
        df.reset_index(drop = "index", inplace = True)

        return df

    if df_type.lower() == "times":
        # Reformatting the columns
        df = reformat_columns(df, relevant_col_idx, "times")

        # Transforming the machine names to ASCII characters.
        df["station"] = format_machine_names(df, "station")

        # Creating a dataframe with the assembly times
        montaj_df = df[(df["station"] == "BANT") | (df["station"] == "LOOP")]

        # Creating a dataframe with the glass bonding
        cmy_df = df[df["station"] == "CAM_YAPISTIRMA"]

        # Dropping the assembly times from the original times dataframe and resetting the index
        df.drop(df[(df["station"] == "BANT") |
                   (df["station"] == "LOOP") |
                   (df["station"] == "CAM_YAPISTIRMA") |
                   (df["part_no"].apply(lambda x: len(x)) == 13)].index, inplace = True)

        # Grouping by the times of the parts that has multiple times in the same work station
        df = df.groupby(["part_no", "station"], as_index = False).agg({"cycle_times": sum, "prep_times": max})
        df.drop(df[df["part_no"].duplicated(keep = "last")].index, inplace = True)

        # Filling the NA codes with 0s (0 codes come from the items that have no time info)
        df[["prep_times", "cycle_times"]] = df[["prep_times", "cycle_times"]].fillna(0)

        # Resetting the index
        df.reset_index(drop = "index", inplace = True)

        return df, montaj_df, cmy_df

    if df_type.lower() == "merged":
        df["station"] = level_lookup(df, "level", "station")
        df["cycle_times"] = level_lookup(df, "level", "cycle_times")
        df["prep_times"] = level_lookup(df, "level", "prep_times")

        df.loc[df.level == 1, ["station", "cycle_times", "prep_times"]] = \
            pd.merge(df["product_no"], assembly_df[["part_no", "station", "cycle_times", "prep_times"]], "left", left_on = "product_no",
                     right_on = "part_no")[["station", "cycle_times", "prep_times"]]

        missing_dict = missing_values_df(df)
        # Continue from here.

        return df


def missing_values_df(df):
    missing_parts = df[df.station.isna()].part_no.apply(lambda x: x.split(".")[0]).unique()
    missing_dict = {}
    for items in missing_parts:
        temp_station = df[df.part_no.apply(lambda x: x.split(".")[0]) == items].station.mode()
        if temp_station.shape == (0,):
            temp_station = np.nan
        else:
            temp_station = temp_station[0]
        temp_cycle = df[df.part_no.apply(lambda x: x.split(".")[0]) == items].cycle_times.mean()
        temp_prep = df[df.part_no.apply(lambda x: x.split(".")[0]) == items].prep_times.mean()
        missing_dict[items] = [temp_station, temp_cycle, temp_prep]
    return missing_dict


def format_word(word, letter_dict):
    new_word = ""
    for i in range(len(word)):
        if word[i] in letter_dict.keys():
            new_word += letter_dict[word[i]].upper()
            continue
        new_word += word[i].upper()
    return new_word


def format_machine_names(df, column):
    turkish_char_dict = {"ı": "i", "ğ": "g", "Ğ": "G", "ü": "u", "Ü": "U", "Ş": "s", "-": "_",
                         "ş": "s", "İ": "I", "Ö": "O", "ö": "o", "Ç": "C", "ç": "c", " ": "_", "/": "_"}
    machine_dict = {}
    for item in list(set(df[column])):
        machine_dict[item] = (format_word(item, turkish_char_dict))
    return [machine_dict[x] for x in df[column]]


def merge_bom_and_times(df_bom, df_times):
    merged_df = pd.merge(left = df_bom, right = df_times, how = "left", on = "part_no")
    merged_df = merged_df.reindex(
        columns = list(merged_df.columns[0:4]) + list(merged_df.columns[5:]) + list(merged_df.columns[4:5]))
    return merged_df


"""
Eski hali ile yapılmış şeyler, denemek için yapıldı bu kısım
Çalıştığından emin olunduktan sonra eski haliyle değiştirilecek
"""


def create_route_matrix(merged_df, bom_df, times_df, montaj_df, inputs_df):
    for i_row in reversed(range(len(inputs_df))):
        w_col = 1
        bom_row = get_level_cross_reference(bom_df, inputs_df.at[i_row, "Ürün Kodu"],
                                            inputs_df.at[i_row, "Bileşen Kodu"])
        bom_next_input = next_input(bom_df, bom_row)
        while (bom_row > bom_next_input) | (bom_row == 0):
            if (not is_input_end(bom_df, bom_row)) | \
                    (find_predecessor_level(bom_df, bom_row) == 1):
                if isinstance(inputs_df.at[i_row, w_col], float):
                    inputs_df.at[i_row, w_col] = get_is_merkezi(times_df, bom_df.at[bom_row - 1, "Bileşen Kodu"])
                    w_col += 1
                    bom_row -= 1
                else:
                    w_col += 1
                    bom_row -= 1
            elif (is_input_end(bom_df, bom_row)) & (get_level(bom_df, bom_row) == 1):
                try:
                    mntj_brm = montaj_df.at[int(bom_df.at[bom_row, "Ürün Kodu"][:4]), "MONTAJ BİRİMİ"]
                except KeyError:
                    mntj_brm = "NI_MNTJ"
                inputs_df.at[i_row, w_col] = mntj_brm
                if bom_row == 0:
                    break
                else:
                    bom_row -= 1
            else:
                inputs_df.at[i_row, w_col] = "BIR_" + get_is_merkezi(times_df, bom_df.at[
                    find_one_upper_level(bom_df, bom_row), "Bileşen Kodu"])
                inputs_df.at[i_row - how_many_inputs_above(bom_df, bom_row),
                             find_last_level(bom_df, find_one_upper_level(bom_df, bom_row)) -
                             get_level(bom_df, bom_row) + 1] = \
                    "BIR_" + get_is_merkezi(times_df, bom_df.at[find_one_upper_level(bom_df, bom_row), "Bileşen Kodu"])
                bom_row -= 1
    return inputs_df
