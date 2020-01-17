import pandas as pd
import numpy as np
from openpyxl import *


def level_lookup(df_name, level_col, lookup_col, level_diff):
    dummies = pd.get_dummies(df_name[level_col])

    idx = dummies.index.to_series()
    last_index = dummies.apply(lambda col: idx.where(col != 0, np.nan).fillna(method = "ffill"))
    last_index[0] = np.nan

    idx = last_index.lookup(last_index.index, df_name[level_col] - level_diff)
    return pd.DataFrame({lookup_col[i]: df_name.reindex(idx)[lookup_col[i]].values for i in range(len(lookup_col))}, index = df_name.index)


def get_excel_sheet_names(directory):
    wb = load_workbook(directory, read_only = True)
    return wb.sheetnames


def read_excel_sheet(directory, name_of_sheet="no_name"):
    if name_of_sheet == "no_name":
        wb = load_workbook(directory, read_only = True)
        sheet = wb[wb.sheetnames[0]]
        df = pd.DataFrame(sheet.values)
        df.columns = df.iloc[0, :]
        df = df.drop(0)
        return df
    wb = load_workbook(directory, read_only = True)
    sheet = wb[name_of_sheet]
    df = pd.DataFrame(sheet.values)
    df.columns = df.iloc[0, :]
    df = df.drop(0)
    return df


def arrange_df(df, relevant_col_idx, df_type, items_to_delete_dir="no_dir"):
    if df_type.lower() == "bom":
        df["Seviye"] = [int(x[-1]) for x in df[df.columns[relevant_col_idx][2]]]
        relevant_col_idx[2] = len(df.columns) - 1
        relevant_col_names = df.columns[relevant_col_idx]
        cols_to_drop = list(set(df.columns) - set(df.columns[relevant_col_idx]))
        df.drop(cols_to_drop, axis = 1, inplace = True)
        df = df.reindex(columns = relevant_col_names)
        df.columns = ["product_no", "part_no", "level", "amount", "explanation"]
        # s = (bom_df[bom_df.columns[2]].ne(bom_df[bom_df.columns[2]].shift(1, fill_value=1)) |
        # bom_df[bom_df.columns[2]].ne(1)).cumsum()
        # bom_df = bom_df.groupby(s).tail(1)
        if df.at[df.index[-1], df.columns[2]] == 1:
            df.drop(df.index[-1], inplace = True)
        df.reset_index(drop = "index", inplace = True)
        items_to_delete = read_excel_sheet(items_to_delete_dir)
        s = [str(x).split(".")[0] not in list(items_to_delete["Kalacaklar"]) for x in df["part_no"]]
        df = df.drop(np.where(np.asarray(s))[0])
        s = (df[df.columns[2]].ne(df[df.columns[2]].shift(1, fill_value = 1)) | df[
            df.columns[2]].ne(1)).cumsum()
        df = df.groupby(s).tail(1)
    if df_type.lower() == "times":
        relevant_col_idx = [3, 25, 29, 30]
        relevant_col_names = df.columns[relevant_col_idx]
        cols_to_drop = list(set(df.columns) - set(df.columns[relevant_col_idx]))
        df.drop(cols_to_drop, axis = 1, inplace = True)
        df = df.reindex(columns = relevant_col_names)
        df.columns = ["part_no", "station", "cycle_times", "prep_times"]
        df = df.groupby(["part_no", "station"], as_index = False).agg(
            {"cycle_times": sum, "prep_times": max})
        df[["prep_times", "cycle_times"]] = df[["prep_times", "cycle_times"]].fillna(0)
    return df


def format_word(word, letter_dict):
    new_word = ""
    for i in range(len(word)):
        if word[i] in letter_dict.keys():
            new_word += letter_dict[word[i]].upper()
            continue
        new_word += word[i].upper()
    return new_word


def format_machine_names(df, column):
    turkish_char_dict = {"ı": "i", "ğ": "g", "Ğ": "G", "ü": "u", "Ü": "U", "Ş": "s", "-": "_",
                         "ş": "s", "İ": "I", "Ö": "O", "ö": "o", "Ç": "C", "ç": "c", " ": "_", "/": "_"}
    machine_dict = {}
    for item in list(set(df[column])):
        machine_dict[item] = (format_word(item, turkish_char_dict))
    return [machine_dict[x] for x in df[column]]


def merge_times_and_bom(df_bom, df_times):
    merged_df = pd.merge(left = df_bom, right = df_times, how = "left", on = "part_no")
    merged_df = merged_df.reindex(
        columns = list(merged_df.columns[0:4]) + list(merged_df.columns[5:]) + list(merged_df.columns[4:5]))
    merged_df["station"] = level_lookup(merged_df, "level", "station", 1)
    merged_df["cycle_times"] = level_lookup(merged_df, "level", "cycle_time", 1)
    merged_df["prep_times"] = level_lookup(merged_df, "level", "prep_times", 1)
    return merged_df

''' 
Eski hali ile yapılmış şeyler, denemek için yapıldı bu kısım
Çalıştığından emin olunduktan sonra eski haliyle değiştirilecek
'''


def create_route_matrix(merged_df, bom_df, times_df, montaj_df, inputs_df):
    for i_row in reversed(range(len(inputs_df))):
        w_col = 1
        bom_row = get_level_cross_reference(bom_df, inputs_df.at[i_row, "Ürün Kodu"], inputs_df.at[i_row, "Bileşen Kodu"])
        bom_next_input = next_input(bom_df, bom_row)
        while (bom_row > bom_next_input) | (bom_row == 0):
            if (not is_input_end(bom_df, bom_row)) | \
                    (find_predecessor_level(bom_df, bom_row) == 1):
                if isinstance(inputs_df.at[i_row, w_col], float):
                    inputs_df.at[i_row, w_col] = get_is_merkezi(times_df, bom_df.at[bom_row - 1, "Bileşen Kodu"])
                    w_col += 1
                    bom_row -= 1
                else:
                    w_col += 1
                    bom_row -= 1
            elif (is_input_end(bom_df, bom_row)) & (get_level(bom_df, bom_row) == 1):
                try:
                    mntj_brm = montaj_df.at[int(bom_df.at[bom_row, "Ürün Kodu"][:4]), "MONTAJ BİRİMİ"]
                except KeyError:
                    mntj_brm = "NI_MNTJ"
                inputs_df.at[i_row, w_col] = mntj_brm
                if bom_row == 0:
                    break
                else:
                    bom_row -= 1
            else:
                inputs_df.at[i_row, w_col] = "BIR_" + get_is_merkezi(times_df, bom_df.at[
                    find_one_upper_level(bom_df, bom_row), "Bileşen Kodu"])
                inputs_df.at[i_row - how_many_inputs_above(bom_df, bom_row),
                          find_last_level(bom_df, find_one_upper_level(bom_df, bom_row)) -
                          get_level(bom_df, bom_row) + 1] = \
                    "BIR_" + get_is_merkezi(times_df, bom_df.at[find_one_upper_level(bom_df, bom_row), "Bileşen Kodu"])
                bom_row -= 1
    return inputs_df
