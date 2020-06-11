import pandas as pd
from datetime import datetime
import random
from openpyxl import load_workbook
from win32com.client import Dispatch
from string import ascii_uppercase as auc
from backend.utils.errors import WrongFileExtensionError, WrongKeywordError

auc = list(auc) + [x + y for x in auc for y in auc]


def get_excel_sheet_names(directory) -> list:
    try:
        wb = load_workbook(directory, read_only=True)
        return wb.sheetnames
    except Exception as e:
        print(e)


def load_file(file_dir, sheet_name=None) -> pd.DataFrame:
    """
    :param sheet_name:
    :param file_dir: The directory of the file that is going to be read.
    :return:
    """
    if file_dir.split(".")[-1].lower() == "csv":
        return pd.read_csv(file_dir, low_memory=False)
    elif (file_dir.split(".")[-1].lower() == "xls") | (file_dir.split(".")[-1].lower() == "xlsx"):
        if sheet_name is None:
            sheet_names = get_excel_sheet_names(file_dir)
            if len(sheet_names) == 1:
                return pd.read_excel(file_dir)
            elif len(sheet_names) > 1:
                print("Which sheet would you like the program to read?:")
                for sn in sheet_names:
                    print(f"[{sheet_names.index(sn) + 1}] - {sn}")
                sheet_choice = int(input("Sheet of choice: "))
                wb = load_workbook(file_dir, read_only=True)
                sheet = wb[sheet_names[sheet_choice - 1]]
                df = pd.DataFrame(sheet.values)
                df.columns = df.iloc[0, :]
                df = df.drop(0)
                df.reset_index(drop=True, inplace=True)
                return df
        else:
            wb = load_workbook(file_dir, read_only=True)
            return pd.DataFrame(wb[sheet_name].values)
    else:
        file_ext = file_dir.split(".")[-1]
        raise WrongFileExtensionError(file_ext)


def load_plan(file_dir):
    plan_df = pd.read_excel(file_dir)[
        ["BANT NO", "ÜRÜNKODU", "Tarih", "PLAN ADET", "Termin Tarihi"]]  # , sheet_name="MartPlan"
    plan_df.columns = ["assembly_unit", "product_no", "start_date", "amount", "due_date"]
    plan_df.drop(plan_df[
                     plan_df.assembly_unit.isnull() | plan_df.product_no.isnull() |
                     plan_df.start_date.isnull() | plan_df.amount.isnull() |
                     plan_df.due_date.isnull()
                     ].index, inplace=True)
    plan_df.sort_values(by="start_date", inplace=True)
    plan_df["due_date"] = plan_df.start_date + pd.to_timedelta([7 for _ in range(plan_df.shape[0])], unit="d")

    plan_df.dropna(inplace=True)
    plan_df.drop(plan_df[pd.to_numeric(plan_df.amount, errors='coerce').isnull()].index, inplace=True)
    plan_df.drop(plan_df[plan_df.due_date.apply(lambda x: type(x)) == str].index, inplace=True)
    plan_df.reset_index(drop=True, inplace=True)
    ourtime = [datetime(pd.to_datetime(plan_df.start_date).dt.year.mode()[0],
                        pd.to_datetime(plan_df.start_date).dt.month.mode()[0] + 1, 1) -
               pd.to_timedelta(1, unit="d") if (pd.to_datetime(plan_df.start_date).dt.month.mode()[0] != 12)
               else datetime(pd.to_datetime(plan_df.start_date).dt.year.mode()[0] + 1, 1, 1) - pd.to_timedelta(1,
                                                                                                               unit="d")
               for _ in range(1)][0]
    plan_df.loc[plan_df["due_date"] > ourtime, "due_date"] = ourtime
    return plan_df


def add_plan(base_plan, file_dir):
    new_plan = load_plan(file_dir)
    if base_plan.iloc[:, :-1].equals(new_plan.iloc[:, :-1]):
        return pd.concat([base_plan, new_plan], ignore_index=True, copy=True), True
    else:
        return pd.concat([base_plan, new_plan], ignore_index=True, copy=True), False


def create_xl_file(input_obj, output_path, file_type):
    if file_type == "operational_simulation_model":
        with pd.ExcelWriter(output_path + "OKPB_Input_File.xlsx") as writer:
            try:
                input_obj.legend_table.to_excel(writer, sheet_name="LEGEND")
                input_obj.input_table.to_excel(writer, sheet_name="input_table")
                input_obj.sequence_matrix.to_excel(writer, sheet_name="sequence_matrix")
                input_obj.time_matrix.to_excel(writer, sheet_name="time_matrix")
                input_obj.join_matrix.to_excel(writer, sheet_name="join_matrix")
                input_obj.join_amount_table.to_excel(writer, sheet_name="join_info")
                input_obj.duplication_matrix.to_excel(writer, sheet_name="product_duplication")
                input_obj.set_list_table.to_excel(writer, sheet_name="set_lists")
                input_obj.order_table.to_excel(writer, sheet_name="orders")
                pd.concat([input_obj.schedules[x] for x in input_obj.schedules.keys()],
                          axis=1).to_excel(writer, sheet_name="schedules", index=False, header=False)
            except ValueError as e:
                print("One of the tables type is not dataframe within the create_excel_file function.")
                print(e)
        namerange_xl_files(input_obj, output_path + "OKPB_Input_File.xlsx", file_type)
    elif file_type == "operational_math_model":
        with pd.ExcelWriter(output_path + "OKPM_Parameters.xlsx") as writer:
            try:
                input_obj.product_no_legend.to_excel(writer, sheet_name="product_no_legend")
                input_obj.machine_legend.to_excel(writer, sheet_name="machine_legend")
                input_obj.plan.to_excel(writer, sheet_name="d")
                input_obj.times.to_excel(writer, sheet_name="h")
                input_obj.route_prob.to_excel(writer, sheet_name="k")
                input_obj.machine_cnt.to_excel(writer, sheet_name="f", index=False)
                input_obj.order_averages.to_excel(writer, sheet_name="o", header=False)
                input_obj.shift.to_excel(writer, sheet_name="s", index=False)
                input_obj.cost.to_excel(writer, sheet_name="c", index=False)
                input_obj.setup_time.to_excel(writer, sheet_name="st", index=False)
                input_obj.outsource_perm.to_excel(writer, sheet_name="osp", index=False)
            except ValueError as e:
                print("One of the tables type is not dataframe within the create_excel_file function.")
                print(e)
        input_obj.machine_legend.to_excel(output_path + "OKPM_Results.xlsx", sheet_name="machine_legend")
    # Discontinued
    elif file_type == "tactical_simulation_model":
        with pd.ExcelWriter(output_path) as writer:
            try:
                input_obj.prod_family_legend_table.to_excel(writer, sheet_name="PROD_FAMILY_LEGEND")
                input_obj.machine_legend_table.to_excel(writer, sheet_name="MACHINE_LEGEND")
                input_obj.sequence_list.to_excel(writer, sheet_name="sequence_list")
                input_obj.min_matrix.to_excel(writer, sheet_name="min_matrix")
                input_obj.mean_matrix.to_excel(writer, sheet_name="mean_matrix")
                input_obj.max_matrix.to_excel(writer, sheet_name="max_matrix")
                input_obj.prob_matrix.to_excel(writer, sheet_name="prob_matrix")
                input_obj.set_list_table.to_excel(writer, sheet_name="set_lists")
                input_obj.order_table.to_excel(writer, sheet_name="orders")
            except ValueError as e:
                print("One of the tables type is not dataframe within the create_excel_file function.")
                print(e)
        namerange_xl_files(input_obj, output_path + "OKPM_Results.xlsx", file_type)
    elif file_type == "tactical_math_model":
        with pd.ExcelWriter(output_path + "TKPM_Parameters.xlsx") as writer:
            try:
                input_obj.product_family_legend.to_excel(writer, sheet_name="product_no_legend", header=False)
                input_obj.machine_legend.to_excel(writer, sheet_name="machine_legend", header=False)
                input_obj.forecast.to_excel(writer, sheet_name="d", merge_cells=False, index_label=False)
                input_obj.times.to_excel(writer, sheet_name="h")
                input_obj.route_prob.to_excel(writer, sheet_name="k")
                input_obj.machine_cnt.to_excel(writer, sheet_name="f", index=False)
                input_obj.workdays.to_excel(writer, sheet_name="w", index=False)
                input_obj.cost.to_excel(writer, sheet_name="c", index=False)
                input_obj.machine_price.to_excel(writer, sheet_name="cr", index=False)
                input_obj.average_order.to_excel(writer, sheet_name="o", header=False)
                input_obj.setup_times.to_excel(writer, sheet_name="st", index=False)
                input_obj.outsource_availability.to_excel(writer, sheet_name="osp", index=False)
                input_obj.order_time_parameters.to_excel(writer, sheet_name="el", index=False)
                input_obj.probabilities.to_excel(writer, sheet_name="pr", index=False)
            except ValueError as e:
                print("One of the tables type is not dataframe within the create_excel_file function.")
                print(e)
        input_obj.machine_legend.to_excel(output_path + "TKPM_Results.xlsx", sheet_name="machine_legend")
        if input_obj.forecast_output is not None:
            input_obj.forecast_output.to_excel(output_path + "Talep Tahmini.xlsx", sheet_name="Talep Tahmini",
                                               merge_cells=False)
    else:
        raise WrongKeywordError(file_type)


def namerange_xl_files(input_obj, output_dir, file_type):
    xl = Dispatch("Excel.Application")
    namerange = "={}!${}$2:${}${}"
    if file_type == "operational_simulation_model":
        info_dict = {
            "coordinate_matrix": {"sheet": "set_lists", "col_start": 5, "num_of_cols": 2,
                                  "num_of_rows": input_obj.set_list_table.shape[0]},
            "duplication_matrix": {"sheet": "product_duplication", "col_start": 3, "num_of_cols": 2,
                                   "num_of_rows": input_obj.duplication_matrix.shape[0]},
            "join_matrix": {"sheet": "join_matrix", "col_start": 2, "num_of_cols": input_obj.join_matrix.shape[1],
                            "num_of_rows": input_obj.join_matrix.shape[0]},
            "join_quantity": {"sheet": "join_info", "col_start": 3, "num_of_cols": 1,
                              "num_of_rows": input_obj.join_amount_table.shape[0]},
            "order_matrix": {"sheet": "orders", "col_start": 2, "num_of_cols": 5,
                             "num_of_rows": input_obj.order_table.shape[0]},
            "part_quantity_list": {"sheet": "input_table", "col_start": 3, "num_of_cols": 1,
                                   "num_of_rows": input_obj.input_table.shape[0]},
            "queues_list": {"sheet": "set_lists", "col_start": 3, "num_of_cols": 1,
                            "num_of_rows": input_obj.set_list_table.shape[0]},
            "resources_list": {"sheet": "set_lists", "col_start": 4, "num_of_cols": 1,
                               "num_of_rows": input_obj.set_list_table.shape[0]},
            "sequence_matrix": {"sheet": "sequence_matrix", "col_start": 2,
                                "num_of_cols": input_obj.sequence_matrix.shape[1],
                                "num_of_rows": input_obj.sequence_matrix.shape[0]},
            "stations_list": {"sheet": "set_lists", "col_start": 2, "num_of_cols": 1,
                              "num_of_rows": input_obj.set_list_table.shape[0]},
            "setup_matrix": {"sheet": "set_lists", "col_start": 7, "num_of_cols": 2,
                             "num_of_rows": input_obj.set_list_table.shape[0]},
            "time_matrix": {"sheet": "time_matrix", "col_start": 2, "num_of_cols": input_obj.time_matrix.shape[1],
                            "num_of_rows": input_obj.time_matrix.shape[0]},
        }  # missing: setup_matrix, setups_list

        xl.Workbooks.Open(Filename=output_dir)
        for namerange_keys in info_dict.keys():
            xl.ActiveWorkbook.Names.Add(Name=namerange_keys, RefersTo=namerange.format(
                info_dict[namerange_keys]["sheet"],  # name of the named range
                auc[info_dict[namerange_keys]["col_start"] - 1],  # the letter that is going to be used for the start
                auc[info_dict[namerange_keys]["col_start"] + info_dict[namerange_keys]["num_of_cols"] - 2],
                # letter of end
                info_dict[namerange_keys]["num_of_rows"] + 1))  # number of the ending column
        # Schedules
        schedule_namerange = "={}!${}$1:${}${}"
        schedule_idx = 0
        schedule_rows = input_obj.schedules[list(input_obj.schedules.keys())[0]].shape[0]
        for schedule_machine in input_obj.schedules.keys():
            xl.ActiveWorkbook.Names.Add(Name="schedule_" + schedule_machine, RefersTo=schedule_namerange.format(
                "schedules",  # name of the named range
                auc[2 * schedule_idx],  # the letter that is going to be used for the start
                auc[2 * schedule_idx + 1],  # letter of end
                schedule_rows))  # number of the ending column
            schedule_idx += 1
    elif file_type == "tactical_simulation_model":
        info_dict = {
            "coordinate_matrix": {"sheet": "set_lists", "col_start": 5, "num_of_cols": 2,
                                  "num_of_rows": input_obj.set_list_table.shape[0]},
            "order_matrix": {"sheet": "orders", "col_start": 2, "num_of_cols": 5,
                             "num_of_rows": input_obj.order_table.shape[0]},
            "queues_list": {"sheet": "set_lists", "col_start": 3, "num_of_cols": 1,
                            "num_of_rows": input_obj.set_list_table.shape[0]},
            "resources_list": {"sheet": "set_lists", "col_start": 4, "num_of_cols": 1,
                               "num_of_rows": input_obj.set_list_table.shape[0]},
            "sequence_list": {"sheet": "sequence_list", "col_start": 2, "num_of_cols": input_obj.sequence_list.shape[1],
                              "num_of_rows": input_obj.sequence_list.shape[0]},
            "stations_list": {"sheet": "set_lists", "col_start": 2, "num_of_cols": 1,
                              "num_of_rows": input_obj.set_list_table.shape[0]},
            "setup_matrix": {"sheet": "set_lists", "col_start": 7, "num_of_cols": 2,
                             "num_of_rows": input_obj.set_list_table.shape[0]},
            "min_matrix": {"sheet": "min_matrix", "col_start": 2, "num_of_cols": input_obj.min_matrix.shape[1],
                           "num_of_rows": input_obj.min_matrix.shape[0]},
            "mean_matrix": {"sheet": "mean_matrix", "col_start": 2, "num_of_cols": input_obj.mean_matrix.shape[1],
                            "num_of_rows": input_obj.mean_matrix.shape[0]},
            "max_matrix": {"sheet": "max_matrix", "col_start": 2, "num_of_cols": input_obj.max_matrix.shape[1],
                           "num_of_rows": input_obj.max_matrix.shape[0]},
            "prob_matrix": {"sheet": "prob_matrix", "col_start": 2, "num_of_cols": input_obj.prob_matrix.shape[1],
                            "num_of_rows": input_obj.prob_matrix.shape[0]},
        }  # missing: setup_matrix, setups_list

        xl.Workbooks.Open(Filename=output_dir)
        for namerange_keys in info_dict.keys():
            xl.ActiveWorkbook.Names.Add(Name=namerange_keys, RefersTo=namerange.format(
                info_dict[namerange_keys]["sheet"],  # name of the named range
                auc[info_dict[namerange_keys]["col_start"] - 1],  # the letter that is going to be used for the start
                auc[info_dict[namerange_keys]["col_start"] + info_dict[namerange_keys]["num_of_cols"] - 2],
                # letter of end
                info_dict[namerange_keys]["num_of_rows"] + 1))  # number of the ending column
    else:
        raise WrongKeywordError(file_type)
    xl.Workbooks(1).Close(SaveChanges=1)
    xl.Application.Quit()


if __name__ == "__main__":
    a = load_plan("C://sl_data//xlsx//planlar.xlsx")
